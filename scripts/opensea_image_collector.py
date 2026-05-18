#!/usr/bin/env python3
"""Collect NFT collection images from workbook contract metadata.

The professor's workbook gives collection names, supplies, OpenSea URLs, and
contract addresses. This script resolves token metadata through public Ethereum
RPC where possible, downloads each image, converts it to a real JPEG, and writes
a manifest row for every attempted token.
"""

from __future__ import annotations

import argparse
import base64
import csv
import io
import json
import re
import zipfile
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import unquote_to_bytes, urlparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from PIL import Image, ImageFile, ImageOps
from web3 import Web3


ERC721_METADATA_ABI = [
    {
        "constant": True,
        "inputs": [{"name": "tokenId", "type": "uint256"}],
        "name": "tokenURI",
        "outputs": [{"name": "", "type": "string"}],
        "payable": False,
        "stateMutability": "view",
        "type": "function",
    }
]

IPFS_GATEWAYS = [
    "https://ipfs.io/ipfs/{path}",
    "https://dweb.link/ipfs/{path}",
    "https://nftstorage.link/ipfs/{path}",
    "https://ipfs.filebase.io/ipfs/{path}",
    "https://gateway.pinata.cloud/ipfs/{path}",
]

KNOWN_TOKEN_URI_TEMPLATES = {
    # BAYC metadata is immutable IPFS JSON by token id. Avoiding one RPC call
    # per token cuts the full 10k collection from hours to minutes.
    "bored-ape-yacht-club": "ipfs://QmeSjSinHpPnmXmspMjwiXyN6zS4E9zccariGR3jxcaWtq/{token_id}",
    "mutant-ape-yacht-club": "https://boredapeyachtclub.com/api/mutants/{token_id}",
    "meebits": "https://meebits.larvalabs.com/meebit/{token_id}",
    "azuki": "ipfs://QmZcH4YvBVVRJtdn4RdbaqgspFU8gH6P9vomDpBVpAL3u4/{token_id}",
    "doodles": "ipfs://QmPMc4tcBsMqLRuCQtPmPe84bpSjrC3Ky7t3JWuHXYB4aS/{token_id}",
    "world-of-women": "ipfs://QmTNBQDbggLZdKF1fRgWnXsnRikd52zL5ciNu769g9JoUP/{token_id}",
    "cool-cats-nft": "https://api.coolcatsnft.com/cat/{token_id}",
    "clone-x": "https://ohm647fhcdf3f6547mcreqj2pgfdzxba7q54ugunppzk3maqy2ma.arweave.net/NGdLvqHyCPyfGzcODgHhOjenXajH--m6mpkp1JLY8M0/{token_id}",
    "cryptoskulls": "ipfs://QmNxWK1AX55kX6oxsiJosu8zQq86geGKj6ApQN1U2hptcL/{token_id}",
    "pudgy-penguins": "ipfs://bafybeibc5sgo2plmjkq2tzmhrn54bk3crhnc23zd2msg4ea7a4pxrkgfna/{token_id}",
    "mooncats": "https://api.mooncat.community/traits/{token_id}",
}

KNOWN_TOKEN_URI_FALLBACK_TEMPLATES = {
    "clone-x": ["https://clonex-assets.rtfkt.com/{token_id}"],
}


@dataclass(frozen=True)
class Collection:
    name: str
    slug: str
    supply: int
    opensea_url: str
    contract_address: str


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "collection"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        default="../OpenSea/List of NFT Collections.xlsx",
        help="Path to List of NFT Collections.xlsx",
    )
    parser.add_argument(
        "--out-dir",
        default="data_lake/opensea",
        help="Output root inside Sharpe-Renaissance",
    )
    parser.add_argument(
        "--rpc-url",
        default="https://ethereum-rpc.publicnode.com",
        help="Ethereum JSON-RPC endpoint for tokenURI calls",
    )
    parser.add_argument(
        "--collections",
        default="",
        help="Comma-separated collection names/slugs. Empty means all workbook rows.",
    )
    parser.add_argument("--start", type=int, default=0, help="First token id to try")
    parser.add_argument("--limit", type=int, default=5, help="Max token ids per collection")
    parser.add_argument(
        "--token-id-file",
        default="",
        help="Optional newline/CSV file of token ids to collect instead of start/limit range.",
    )
    parser.add_argument("--sleep", type=float, default=0.25, help="Delay between tokens")
    parser.add_argument("--workers", type=int, default=1, help="Parallel download workers")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing JPG files")
    parser.add_argument(
        "--save-metadata",
        action="store_true",
        help="Write raw token metadata JSON files next to the image dataset.",
    )
    return parser.parse_args()


def read_collections(path: Path) -> list[Collection]:
    if path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path) as zf:
            matches = [name for name in zf.namelist() if name.lower().endswith(".xlsx")]
            if not matches:
                raise FileNotFoundError(f"no .xlsx workbook found in {path}")
            with zf.open(matches[0]) as fh:
                wb = load_workbook(fh, data_only=True)
                return read_collections_from_workbook(wb)

    wb = load_workbook(path, data_only=True)
    return read_collections_from_workbook(wb)


def read_collections_from_workbook(wb: Any) -> list[Collection]:
    ws = wb["OpenSea"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x or "").strip() for x in rows[0]]
    idx = {name: headers.index(name) for name in headers}

    out: list[Collection] = []
    for row in rows[1:]:
        name = str(row[idx["NFT Collection"]] or "").strip()
        if not name:
            continue
        supply_raw = row[idx["Supply"]]
        supply = int(supply_raw or 0)
        out.append(
            Collection(
                name=name,
                slug=slugify(name),
                supply=supply,
                opensea_url=str(row[idx["OpenSea"]] or "").strip(),
                contract_address=str(row[idx["ContractAddress"]] or "").strip(),
            )
        )
    return out


def read_token_ids(path: Path) -> list[int]:
    token_ids: list[int] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        for part in line.split(","):
            part = part.strip()
            if part:
                token_ids.append(int(part))
    return token_ids


def candidate_urls(uri: str) -> list[str]:
    uri = (uri or "").strip()
    if not uri:
        return []
    if uri.startswith("ipfs://"):
        path = uri.removeprefix("ipfs://").lstrip("/")
        return [gateway.format(path=path) for gateway in IPFS_GATEWAYS]
    if uri.startswith("ar://"):
        path = uri.removeprefix("ar://").lstrip("/")
        return [f"https://arweave.net/{path}"]
    if uri.startswith("http://") or uri.startswith("https://"):
        parsed = urlparse(uri)
        if "/ipfs/" in parsed.path:
            path = parsed.path.split("/ipfs/", 1)[1].lstrip("/")
            urls = [uri]
            urls.extend(gateway.format(path=path) for gateway in IPFS_GATEWAYS)
            return list(dict.fromkeys(urls))
        if parsed.netloc.endswith("arweave.net") and parsed.path.lower().endswith((".png", ".jpg", ".jpeg")):
            return [uri, uri.rsplit(".", 1)[0]]
        return [uri]
    return [uri]


def parse_data_uri(uri: str) -> tuple[bytes, str]:
    header, payload = uri.split(",", 1)
    ctype = header.removeprefix("data:").split(";", 1)[0] or "text/plain"
    if ";base64" in header:
        return base64.b64decode(payload), ctype
    return unquote_to_bytes(payload), ctype


def fetch_url(
    session: requests.Session,
    uri: str,
    *,
    expect_json: bool = False,
    attempts: int = 5,
) -> tuple[str, bytes, str]:
    if uri.startswith("data:"):
        raw, ctype = parse_data_uri(uri)
        if expect_json and "json" not in ctype.lower() and not raw.lstrip().startswith(b"{"):
            raise RuntimeError(f"non-json data URI {ctype}")
        return "data:", raw, ctype

    last_error = ""
    for url in candidate_urls(uri):
        for attempt in range(1, attempts + 1):
            try:
                resp = session.get(url, timeout=(10, 30))
                ctype = resp.headers.get("content-type", "")
                if resp.ok:
                    if expect_json and "json" not in ctype.lower() and not resp.text.lstrip().startswith("{"):
                        last_error = f"non-json response {resp.status_code} {ctype} {url}"
                        break
                    return url, resp.content, ctype
                last_error = f"http {resp.status_code} {url}"
                if resp.status_code in {408, 425, 429, 500, 502, 503, 504} and attempt < attempts:
                    retry_after = resp.headers.get("retry-after")
                    delay = float(retry_after) if retry_after and retry_after.isdigit() else min(8.0, 0.75 * 2 ** (attempt - 1))
                    time.sleep(delay)
                    continue
                break
            except Exception as exc:  # noqa: BLE001 - manifest should capture all network failures.
                last_error = f"{type(exc).__name__}: {exc}"
                if attempt < attempts:
                    time.sleep(min(8.0, 0.75 * 2 ** (attempt - 1)))
                    continue
                break
    raise RuntimeError(last_error or f"no candidate URL for {uri}")


def convert_to_jpg(raw: bytes, out_path: Path) -> tuple[int, int, str]:
    try:
        img_context = Image.open(io.BytesIO(raw))
        img_context.load()
    except OSError as exc:
        if "truncated" not in str(exc).lower():
            raise
        previous = ImageFile.LOAD_TRUNCATED_IMAGES
        ImageFile.LOAD_TRUNCATED_IMAGES = True
        try:
            img_context = Image.open(io.BytesIO(raw))
            img_context.load()
        finally:
            ImageFile.LOAD_TRUNCATED_IMAGES = previous

    with img_context as img:
        source_format = img.format or ""
        img = ImageOps.exif_transpose(img)
        if img.mode in {"RGBA", "LA"}:
            bg = Image.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.getchannel("A"))
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        img.save(out_path, "JPEG", quality=95, optimize=True)
        return img.width, img.height, source_format


def fetch_and_convert_image(
    session: requests.Session,
    image_uri: str,
    out_path: Path,
    *,
    attempts: int = 2,
) -> tuple[str, int, int, str]:
    if image_uri.startswith("data:"):
        raw, _ctype = parse_data_uri(image_uri)
        width, height, source_format = convert_to_jpg(raw, out_path)
        return "data:", width, height, source_format

    last_error = ""
    for url in candidate_urls(image_uri):
        for attempt in range(1, attempts + 1):
            try:
                resp = session.get(url, timeout=(5, 15))
                if resp.ok:
                    try:
                        width, height, source_format = convert_to_jpg(resp.content, out_path)
                        return url, width, height, source_format
                    except Exception as exc:  # noqa: BLE001 - try alternate image candidates.
                        last_error = f"{type(exc).__name__}: {exc} {url}"
                        break
                last_error = f"http {resp.status_code} {url}"
                if resp.status_code in {408, 425, 429, 500, 502, 503, 504} and attempt < attempts:
                    retry_after = resp.headers.get("retry-after")
                    delay = float(retry_after) if retry_after and retry_after.isdigit() else min(8.0, 0.75 * 2 ** (attempt - 1))
                    time.sleep(delay)
                    continue
                break
            except Exception as exc:  # noqa: BLE001 - manifest should capture all network failures.
                last_error = f"{type(exc).__name__}: {exc}"
                if attempt < attempts:
                    time.sleep(min(8.0, 0.75 * 2 ** (attempt - 1)))
                    continue
                break
    raise RuntimeError(last_error or f"no image candidate URL for {image_uri}")


def cryptopunks_image_uri(token_id: int) -> str:
    return f"https://www.larvalabs.com/cryptopunks/cryptopunk{token_id}.png"


def short_error(exc: Exception, limit: int = 500) -> str:
    text = f"{type(exc).__name__}: {exc}"
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > limit:
        return text[:limit] + "..."
    return text


def manifest_uri(uri: str, limit: int = 500) -> str:
    uri = uri or ""
    if uri.startswith("data:"):
        header, _sep, payload = uri.partition(",")
        return f"{header},<embedded:{len(payload)} chars>"
    if len(uri) > limit:
        return uri[:limit] + "..."
    return uri


def resolve_token_metadata(
    w3: Web3,
    session: requests.Session,
    collection: Collection,
    token_id: int,
) -> tuple[str, str, dict[str, Any]]:
    if collection.slug == "cryptopunks":
        image_uri = cryptopunks_image_uri(token_id)
        return "", image_uri, {"source": "larvalabs_cryptopunks_image"}

    token_uri_template = KNOWN_TOKEN_URI_TEMPLATES.get(collection.slug)
    if token_uri_template:
        token_uri = token_uri_template.format(token_id=token_id)
    else:
        contract = w3.eth.contract(
            address=Web3.to_checksum_address(collection.contract_address),
            abi=ERC721_METADATA_ABI,
        )
        token_uri = contract.functions.tokenURI(token_id).call()
    try:
        metadata_url, raw, _ctype = fetch_url(session, token_uri, expect_json=True)
    except Exception:
        fallback_templates = KNOWN_TOKEN_URI_FALLBACK_TEMPLATES.get(collection.slug, [])
        if not fallback_templates:
            raise
        last_error = None
        for fallback_template in fallback_templates:
            fallback_uri = fallback_template.format(token_id=token_id)
            try:
                metadata_url, raw, _ctype = fetch_url(session, fallback_uri, expect_json=True)
                break
            except Exception as exc:  # noqa: BLE001 - try all configured metadata fallbacks.
                last_error = exc
        else:
            raise last_error
    metadata = json.loads(raw.decode("utf-8"))
    image_uri = metadata.get("image") or metadata.get("image_url") or metadata.get("image_data")
    if not image_uri:
        raise RuntimeError(f"metadata has no image field: {metadata_url}")
    metadata["_metadata_url"] = metadata_url
    return token_uri, str(image_uri), metadata


def manifest_fields() -> list[str]:
    return [
        "collection",
        "slug",
        "token_id",
        "status",
        "output_path",
        "width",
        "height",
        "source_format",
        "token_uri",
        "metadata_url",
        "image_uri",
        "image_url",
        "metadata_path",
        "name",
        "description",
        "attributes_json",
        "error",
    ]


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook).expanduser().resolve()
    out_root = Path(args.out_dir).expanduser().resolve()
    image_root = out_root / "images"
    metadata_root = out_root / "metadata"
    manifest_path = out_root / "manifests" / "download_manifest.csv"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    collections = read_collections(workbook)
    wanted = {slugify(x) for x in args.collections.split(",") if x.strip()}
    if wanted:
        collections = [c for c in collections if c.slug in wanted or slugify(c.name) in wanted]

    probe_w3 = Web3(Web3.HTTPProvider(args.rpc_url, request_kwargs={"timeout": 30}))
    if not probe_w3.is_connected():
        raise SystemExit(f"RPC not reachable: {args.rpc_url}")

    local = threading.local()
    token_id_override = read_token_ids(Path(args.token_id_file).expanduser().resolve()) if args.token_id_file else []

    def worker_session() -> requests.Session:
        session = getattr(local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update(
                {
                    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/121 Safari/537.36",
                    "Accept": "application/json,image/avif,image/webp,image/png,image/jpeg,*/*",
                }
            )
            local.session = session
        return session

    def worker_w3() -> Web3:
        w3 = getattr(local, "w3", None)
        if w3 is None:
            w3 = Web3(Web3.HTTPProvider(args.rpc_url, request_kwargs={"timeout": 30}))
            local.w3 = w3
        return w3

    def process_token(collection: Collection, token_id: int) -> dict[str, Any]:
        rel_out = Path(collection.slug) / f"{token_id}.jpg"
        out_path = image_root / rel_out
        row: dict[str, Any] = {
            "collection": collection.name,
            "slug": collection.slug,
            "token_id": token_id,
            "status": "",
            "output_path": str(out_path.relative_to(out_root)),
            "width": "",
            "height": "",
            "source_format": "",
            "token_uri": "",
            "metadata_url": "",
            "image_uri": "",
            "image_url": "",
            "metadata_path": "",
            "name": "",
            "description": "",
            "attributes_json": "",
            "error": "",
        }
        try:
            if out_path.exists() and not args.overwrite:
                with Image.open(out_path) as img:
                    row.update(
                        {
                            "status": "exists",
                            "width": img.width,
                            "height": img.height,
                            "source_format": img.format or "JPEG",
                        }
                    )
                return row

            token_uri, image_uri, metadata = resolve_token_metadata(
                worker_w3(),
                worker_session(),
                collection,
                token_id,
            )
            metadata_rel = Path(collection.slug) / f"{token_id}.json"
            if args.save_metadata:
                metadata_path = metadata_root / metadata_rel
                metadata_path.parent.mkdir(parents=True, exist_ok=True)
                metadata_path.write_text(
                    json.dumps(metadata, ensure_ascii=False, indent=2, sort_keys=True),
                    encoding="utf-8",
                )
            image_url, width, height, source_format = fetch_and_convert_image(worker_session(), image_uri, out_path)
            row.update(
                {
                    "status": "ok",
                    "width": width,
                    "height": height,
                    "source_format": source_format,
                    "token_uri": manifest_uri(token_uri),
                    "metadata_url": manifest_uri(metadata.get("_metadata_url", "")),
                    "image_uri": manifest_uri(image_uri),
                    "image_url": manifest_uri(image_url),
                    "metadata_path": str((Path("metadata") / metadata_rel).as_posix()) if args.save_metadata else "",
                    "name": str(metadata.get("name", "")),
                    "description": str(metadata.get("description", "")),
                    "attributes_json": json.dumps(metadata.get("attributes", []), ensure_ascii=False, separators=(",", ":")),
                }
            )
            return row
        except Exception as exc:  # noqa: BLE001 - errors are the audit trail.
            row.update({"status": "error", "error": short_error(exc)})
            return row

    exists = manifest_path.exists()
    with manifest_path.open("a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=manifest_fields())
        if not exists:
            writer.writeheader()

        ok = 0
        failed = 0
        for collection in collections:
            if token_id_override:
                token_ids = token_id_override
            else:
                max_token = min(collection.supply, args.start + args.limit)
                token_ids = range(args.start, max_token)
            workers = max(1, int(args.workers or 1))
            if workers == 1:
                row_iter = (process_token(collection, token_id) for token_id in token_ids)
            else:
                executor = ThreadPoolExecutor(max_workers=workers)
                futures = [executor.submit(process_token, collection, token_id) for token_id in token_ids]
                row_iter = (future.result() for future in as_completed(futures))

            for row in row_iter:
                if row["status"] in {"ok", "exists"}:
                    ok += 1
                    print(f"{row['status']} {collection.slug} #{row['token_id']} -> {row['output_path']}", flush=True)
                else:
                    failed += 1
                    print(f"error {collection.slug} #{row['token_id']}: {row['error']}", flush=True)
                writer.writerow(row)
                fh.flush()
                if args.sleep:
                    time.sleep(args.sleep)
            if workers > 1:
                executor.shutdown(wait=True)

    print(f"done collections={len(collections)} ok={ok} failed={failed} manifest={manifest_path}")
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
